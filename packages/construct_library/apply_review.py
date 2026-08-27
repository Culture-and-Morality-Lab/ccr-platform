#!/usr/bin/env python3
"""Apply the 2026-08-25 construct library review to the YAML library (spec 0007).

Noor Skhiri reviewed all 525 items across the 94 constructs, checking wording,
reverse keying, subscale grouping, and citation for each. This script turns that
review file into library changes, deterministically and re-runnably.

What lands where is decided by the item hash (see validate_constructs.item_hash),
which covers language, version, item text, item order, and reverse flags:

  * a change to any of those creates a NEW version file (append-only, spec 0004) -
    the old version file is left byte-identical so runs that used it still resolve;
  * everything else (verification_status, citation, source_url, questionnaire,
    review provenance) is outside the hash and is edited in place.

Corrected item text and replacement URLs are hand-specified in the tables below
rather than parsed out of the reviewer's free-text notes: the notes mix the
correction with commentary, and a bad regex here would silently corrupt a
validated scale. The reverse-flag flips ARE read from the sheet, because that
column is a clean Y/N.

Usage:
    python packages/construct_library/apply_review.py            # dry run
    python packages/construct_library/apply_review.py --write    # rewrite YAML
Then: python packages/construct_library/validate_constructs.py
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import yaml

HERE = Path(__file__).parent
CONSTRUCTS = HERE / "constructs"
REVIEW_XLSX = HERE / "reviews" / "CCR_construct_library_review_2026-08-25.xlsx"

REVIEWER = "Noor Skhiri"
REVIEW_DATE = "2026-08-25"

# Constructs whose remaining findings need a PI decision (spec 0007 Non-goals).
# They still receive their uncontroversial reverse-flag fixes; they just do not
# graduate to `verified`.
PENDING = {
    "ipip_50_item_big_five_factor_markers_agreeableness",
    "ipip_50_item_big_five_factor_markers_conscientiousness",
    "ipip_50_item_big_five_factor_markers_emotional_stability_neuroticism",
    "ipip_50_item_big_five_factor_markers_extraversion",
    "ipip_50_item_big_five_factor_markers_intellect_imagination",
    "k10",
}
PENDING_NOTE = {
    "k10": (
        "Items are stored as bare fragments; reviewer flags that each should carry the "
        "source stem 'During the last 30 days, about how often did you feel'. Pending PI "
        "decision (spec 0007)."
    ),
}
IPIP_NOTE = (
    "Reviewer confirms wording matches the source apart from the leading 'I' that CCR "
    "adds to IPIP item stems; whether to keep that prefix is pending a PI decision "
    "(spec 0007). Reverse-scoring corrections from the review are applied."
)

# item_id -> corrected text, transcribed from the reviewer's notes.
WORDING_FIX = {
    "adult_hope_scale_pathways_6":
        "I can think of many ways to get the things in life that are important to me.",
    "bas_2_8":
        "My behavior reveals my positive attitude toward my body; for example, I walk "
        "holding my head high and smiling.",
    "cage_questionnaire_1":
        "Have you ever felt you should Cut down on your drinking?",
    "hi_3":
        'I often do "my own thing."',
    "mfq_2_proportionality_3":
        "I think people who are more hardworking should end up with more money.",
    "mspss_significant_other_2":
        "There is a special person with whom I can share joys and sorrows.",
    "shs_2":
        "Compared to most of my peers, I consider myself:",
    # The reviewer read this off the questionnaire printed in Graham et al.
    # (2011), which is the paper this construct cites, and the paper does say
    # "from others" (twice). The moralfoundations.org MFQ30 handout says "than
    # others"; the cited paper wins.
    "fair_3":
        "Whether or not some people were treated differently from others.",
    # NOT applied: team_psychological_safety_scale_4. The reviewer proposed "risk
    # in this team" from a third-party questionnaire because she could not reach
    # Edmondson (1999) through the library, and said to keep the original if the
    # paper disagreed. It does: the recorded Edmondson PDF reads "risk on this
    # team", so the stored wording stands.
}

# Item ORDER corrections. Order feeds item_hash and decides which item each
# sim_item_N export column refers to, so a construct whose stored order does not
# match its own cited source mislabels every per-item column. Values are the
# construct's item_ids in the source's printed order; ids are renumbered by
# position afterwards so item_id once again means "source item number".
# Both were checked against the source_url recorded on the construct.
ITEM_ORDER = {
    # socy.umd.edu questionnaire image (Morris Rosenberg Foundation): item 1 is
    # "I feel that I'm a person of worth". Wording there matches ours exactly.
    "rses": ["rses_7", "rses_3", "rses_9", "rses_4", "rses_5",
             "rses_10", "rses_1", "rses_8", "rses_6", "rses_2"],
    # emerge.ucsd.edu CBI: the work-related subscale runs 7..13 starting with
    # "Do you feel worn out at the end of the working day?".
    "cbi_work_related_burnout": [
        "cbi_work_related_burnout_10", "cbi_work_related_burnout_11",
        "cbi_work_related_burnout_12", "cbi_work_related_burnout_13",
        "cbi_work_related_burnout_7", "cbi_work_related_burnout_9",
        "cbi_work_related_burnout_8",
    ],
}
# Source item number of the first item in each reordered construct, so renumbered
# ids keep matching the printed questionnaire.
ITEM_ORDER_START = {"rses": 1, "cbi_work_related_burnout": 7}

# One item moves subscale per the SCS-SF coding key. It is renumbered into the
# destination construct's own id convention ({construct}_{source item number});
# it is still SCS-SF item 1, and the review file records where it came from.
SUBSCALE_MOVE = {
    "from": "scs_sf_self_judgment",
    "to": "scs_sf_over_identification",
    "item_id": "scs_sf_self_judgment_1",
    "new_item_id": "scs_sf_over_identification_1",
    "insert_at": 0,
}
SUBSCALE_MOVE_NOTE = (
    "SCS-SF item 1 ('When I fail at something important to me...') moves from "
    "Self-judgment to Over-identification per the source coding key, and is renumbered "
    "to match this construct's item ids (spec 0007)."
)

# Replacement source URLs supplied by the reviewer. PsycNET links arrived with an
# `auth_token` query parameter: that is a per-session credential which expires and
# must not be committed, so it is stripped and the base URL kept (the citation DOI
# stays the durable pointer).
SOURCE_URL_FIX = {
    "cbi_client_related_burnout": "https://emerge.ucsd.edu/r_2qfb6wi4uepyugd/",
    "cbi_personal_burnout": "https://emerge.ucsd.edu/r_2qfb6wi4uepyugd/",
    "cbi_work_related_burnout": "https://emerge.ucsd.edu/r_2qfb6wi4uepyugd/",
    "flourishing_scale": "https://eddiener.com/wp-content/uploads/2024/09/Flourishing-Scale.pdf",
    "grit_s_consistency_of_interests":
        "https://www.dropbox.com/scl/fi/5fw2nbvvswu6jfb/8-item-Grit-4.pdf?rlkey=htfr2ngc17y027uv9ebnj53zm&dl=0",
    "grit_s_perseverance_of_effort":
        "https://www.dropbox.com/scl/fi/5fw2nbvvswu6jfb/8-item-Grit-4.pdf?rlkey=htfr2ngc17y027uv9ebnj53zm&dl=0",
    # team_psychological_safety_scale is deliberately absent: the reviewer's
    # replacement is a third-party questionnaire whose wording disagrees with the
    # cited paper, and the recorded Edmondson PDF still resolves and still
    # contains the scale.
    "dirty_dozen_machiavellianism": "https://psycnet.apa.org/fulltext/2010-10892-021.pdf",
    "dirty_dozen_narcissism": "https://psycnet.apa.org/fulltext/2010-10892-021.pdf",
    "dirty_dozen_psychopathy": "https://psycnet.apa.org/fulltext/2010-10892-021.pdf",
    "lot_r": "https://psycnet.apa.org/fulltext/1995-07978-001.pdf",
    # The sheet went out with no source_url for these two, so the reviewer had
    # nothing to check against. Added afterwards and confirmed item by item.
    # The reviewer read the MFQ off the questionnaire printed at the end of
    # Graham et al. (2011), which is the paper both constructs cite. Pointing at
    # the paper rather than the later MFQ30 handout keeps items, citation, and
    # source in agreement (they differ on one item's wording).
    "mfq_care": "https://www.the-brights.net/morality/statement_1_studies/14.%20Graham,%20J.,%20Nosek,%20B.%20A.,%20Haidt,%20J.,%20Iyer,%20R.,%20Koleva,%20S.,%20&%20Ditto,%20P.%20H.%20(2011).%20Mapping%20the%20moral%20domain.%20Journal%20of%20Personality%20and%20Social%20Psychology,%20101(2),%20366-385.pdf",
    "mfq_fairness": "https://www.the-brights.net/morality/statement_1_studies/14.%20Graham,%20J.,%20Nosek,%20B.%20A.,%20Haidt,%20J.,%20Iyer,%20R.,%20Koleva,%20S.,%20&%20Ditto,%20P.%20H.%20(2011).%20Mapping%20the%20moral%20domain.%20Journal%20of%20Personality%20and%20Social%20Psychology,%20101(2),%20366-385.pdf",
    # Replacements supplied by the reviewer after the review; both verified to
    # resolve and to contain the items. The recorded links were dead (BAS-2 404,
    # CAGE 403 even from a browser user agent).
    "bas_2": "https://emerge.ucsd.edu/r_264jgxeqd35y1ox/",
    "cage_questionnaire": "https://www.uspreventiveservicestaskforce.org/home/getfilebytoken/QETZFZLBsJbtPmVNVCgA7Z",
    "collectivism_horizontal": "https://psycnet.apa.org/fulltext/1997-38342-009.pdf",
    "individualism_horizontal": "https://psycnet.apa.org/fulltext/1997-38342-009.pdf",
}

# Citations the reviewer marked incomplete. Every DOI and page range below appears
# verbatim in her notes; no bibliographic detail is looked up or inferred. The only
# editorial change is expanding the journal abbreviation "JPSP", which the existing
# citation already named.
CITATION_FIX = {
    # Reviewer supplied the full reference including "49(1), 71-75" and the DOI.
    "satisfaction_with_life":
        "Diener, E., Emmons, R. A., Larsen, R. J., & Griffin, S. (1985). The Satisfaction "
        "with Life Scale. Journal of Personality Assessment, 49(1), 71-75. "
        "https://doi.org/10.1207/s15327752jpa4901_13",
    # Reviewer supplied the DOI only, so no page range is added here.
    "collectivism_horizontal":
        "Triandis, H. C., & Gelfand, M. J. (1998). Converging measurement of horizontal and "
        "vertical individualism and collectivism. Journal of Personality and Social "
        "Psychology, 74(1). https://doi.org/10.1037/0022-3514.74.1.118",
    "individualism_horizontal":
        "Triandis, H. C., & Gelfand, M. J. (1998). Converging measurement of horizontal and "
        "vertical individualism and collectivism. Journal of Personality and Social "
        "Psychology, 74(1). https://doi.org/10.1037/0022-3514.74.1.118",
    # Page range and DOI are printed on the paper itself (now the recorded
    # source_url), not looked up.
    "mfq_care":
        "Graham, J., Nosek, B. A., Haidt, J., Iyer, R., Koleva, S., & Ditto, P. H. (2011). "
        "Mapping the moral domain. Journal of Personality and Social Psychology, 101(2), "
        "366-385. https://doi.org/10.1037/a0021847",
    "mfq_fairness":
        "Graham, J., Nosek, B. A., Haidt, J., Iyer, R., Koleva, S., & Ditto, P. H. (2011). "
        "Mapping the moral domain. Journal of Personality and Social Psychology, 101(2), "
        "366-385. https://doi.org/10.1037/a0021847",
}

QUESTIONNAIRE_FIX = {
    "collectivism_horizontal": "Horizontal and Vertical Individualism and Collectivism Scale",
    "individualism_horizontal": "Horizontal and Vertical Individualism and Collectivism Scale",
}

# Notes on constructs that DO reach `verified`. Gaps serious enough to hold a
# construct back live in PENDING_NOTE instead.
REVIEW_NOTE = {
    "lot_r": "The scale's 4 filler items are correctly excluded from this construct; they are "
             "not scored in the source.",
    "mfq_care": "The review sheet shipped without a source_url for this construct. The "
                "questionnaire printed in the cited paper is now recorded and all 4 items "
                "confirmed verbatim against it. The construct uses a subset of the foundation's "
                "items, which the reviewer confirmed as correctly grouped.",
    "mfq_fairness": "The review sheet shipped without a source_url. The reviewer read the MFQ "
                    "off the appendix of the cited paper, which reads 'treated differently from "
                    "others'; her correction is applied and all 4 items confirmed verbatim. The "
                    "later MFQ30 handout says 'than others', so items, citation, and source_url "
                    "now all point at the paper.",
    "team_psychological_safety_scale": "Reviewer proposed 'risk in this team' from a third-party "
                                       "questionnaire because Edmondson (1999) was not reachable "
                                       "through her library, and said to keep the original if the "
                                       "paper disagreed. The recorded Edmondson PDF reads 'risk on "
                                       "this team', so the stored wording stands.",
    "rses": "Item order corrected to the printed order of the questionnaire at the recorded "
            "source_url (Morris Rosenberg Foundation), per the reviewer's mapping; item ids "
            "renumbered to match. Wording and reverse keys there match this construct exactly.",
    "cbi_work_related_burnout": "Item order corrected to the printed order at the recorded "
                                "source_url, per the reviewer's mapping; item ids renumbered to "
                                "match the source's 7 to 13.",
    "bas_2": "Recorded source URL was dead (404). Replacement supplied by the reviewer, and all "
             "10 items confirmed verbatim against it. Her note that BAS-2 has no reverse-scored "
             "items and is scored by averaging all 10 is consistent with the flags here.",
    "cage_questionnaire": "Recorded source URL returned 403. Replacement supplied by the reviewer "
                          "and the wording correction ('ought to' -> 'should') confirmed against "
                          "it. The replacement is a token-style URL and may not be durable.",
}


def load_review() -> dict:
    """Read the reviewer's sheet. Returns {item_id: should_be_reverse_scored}.

    An unanswered cell is an error, not a False: treating a blank as "not
    reverse-scored" would silently un-flag a reversed item, which is a scoring
    change nobody asked for and nothing downstream would surface.
    """
    ws = openpyxl.load_workbook(REVIEW_XLSX, data_only=True)["Items"]
    hdr = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(hdr) if h}
    flags, bad = {}, []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[0]:
            continue
        item_id = str(r[idx["item_id"]]).strip()
        answer = str(r[idx["REVERSE-SCORED? (Y/N)"]] or "").strip().upper()
        if answer not in ("Y", "N"):
            bad.append(f"{item_id}: reverse-scored cell is {answer or 'blank'!r}")
            continue
        flags[item_id] = answer == "Y"
    if bad:
        raise ValueError(
            f"{len(bad)} item(s) have no usable reverse-scoring answer; resolve them with "
            "the reviewer before applying:\n  " + "\n  ".join(bad[:20])
        )
    return flags


def dump(path: Path, data: dict, header: str) -> None:
    body = yaml.safe_dump(data, sort_keys=False, allow_unicode=True, width=100,
                          default_flow_style=False)
    path.write_text(header + body)


def main() -> int:
    write = "--write" in sys.argv
    if not REVIEW_XLSX.exists():
        print(f"Review file not found: {REVIEW_XLSX}")
        return 1

    review_flags = load_review()
    files = sorted(CONSTRUCTS.glob("*.yaml"))
    plan = {"new_version": [], "in_place": [], "unchanged": [], "detail": []}

    for f in files:
        c = yaml.safe_load(f.read_text())
        cid = c["construct_id"]
        if int(c["version"]) != 1:
            continue  # already-versioned output from a previous run

        items = [dict(i) for i in c["items"]]

        # --- item-level changes (these move the hash, so they need a new version)
        changed_items = False
        for it in items:
            iid = it["item_id"]
            if iid in WORDING_FIX and it["text"] != WORDING_FIX[iid]:
                plan["detail"].append(f"  {cid}/{iid}: text -> {WORDING_FIX[iid]!r}")
                it["text"] = WORDING_FIX[iid]
                changed_items = True
            if iid in review_flags and bool(it.get("reverse_scored", False)) != review_flags[iid]:
                arrow = "no -> REVERSE" if review_flags[iid] else "REVERSE -> no"
                plan["detail"].append(f"  {cid}/{iid}: {arrow}")
                it["reverse_scored"] = review_flags[iid]
                changed_items = True

        if cid in ITEM_ORDER:
            wanted = ITEM_ORDER[cid]
            by_id = {i["item_id"]: i for i in items}
            missing = [x for x in wanted if x not in by_id]
            assert not missing and len(wanted) == len(items), (
                f"{cid}: ITEM_ORDER does not cover the construct exactly ({missing=})"
            )
            if [i["item_id"] for i in items] != wanted:
                items = [by_id[x] for x in wanted]
                # ids encode the source's item number, so renumber by position
                start = ITEM_ORDER_START[cid]
                for n, it in enumerate(items, start=start):
                    it["item_id"] = f"{cid}_{n}"
                plan["detail"].append(
                    f"  {cid}: reordered to the source's printed order, ids renumbered "
                    f"{start}..{start + len(items) - 1}"
                )
                changed_items = True

        src, dst = SUBSCALE_MOVE["from"], SUBSCALE_MOVE["to"]
        move_id, new_id = SUBSCALE_MOVE["item_id"], SUBSCALE_MOVE["new_item_id"]
        if cid == src and any(i["item_id"] == move_id for i in items):
            items = [i for i in items if i["item_id"] != move_id]
            changed_items = True
        elif cid == dst and not any(i["item_id"] == new_id for i in items):
            donor = yaml.safe_load((CONSTRUCTS / f"{src}.yaml").read_text())
            incoming = next(dict(i) for i in donor["items"] if i["item_id"] == move_id)
            incoming["reverse_scored"] = review_flags.get(move_id, incoming.get("reverse_scored", False))
            incoming["item_id"] = new_id
            items.insert(SUBSCALE_MOVE["insert_at"], incoming)
            changed_items = True

        # --- metadata (outside the hash: edited in place on whichever file is live)
        live = dict(c)
        if changed_items:
            live["version"] = 2
            live["items"] = items
        status = "needs_verification" if cid in PENDING else "verified"
        live["verification_status"] = status
        if cid in SOURCE_URL_FIX:
            live["source_url"] = SOURCE_URL_FIX[cid]
        if cid in CITATION_FIX:
            live["citation"] = CITATION_FIX[cid]
        if cid in QUESTIONNAIRE_FIX:
            live["questionnaire"] = QUESTIONNAIRE_FIX[cid]
        if "reverse_flags_source" in live or changed_items:
            live["reverse_flags_source"] = f"reviewed_{REVIEW_DATE}_{REVIEWER.lower().replace(' ', '_')}"

        note = REVIEW_NOTE.get(cid) or PENDING_NOTE.get(cid)
        if cid in PENDING and cid.startswith("ipip_"):
            note = IPIP_NOTE
        if cid in (src, dst):
            note = f"{note} {SUBSCALE_MOVE_NOTE}".strip() if note else SUBSCALE_MOVE_NOTE
        review = {"reviewer": REVIEWER, "date": REVIEW_DATE,
                  "outcome": "verified" if status == "verified" else "pending_pi_decision"}
        if note:
            review["notes"] = note
        live["review"] = review
        # keep `items` last for readability
        live["items"] = live.pop("items")

        if changed_items:
            plan["new_version"].append(cid)
            if write:
                # v1 stays byte-identical except for being marked superseded:
                # its items are the pre-review ones and must never read as verified.
                old = dict(c)
                old["verification_status"] = "archived"
                old["review"] = {"reviewer": REVIEWER, "date": REVIEW_DATE,
                                 "outcome": "superseded_by_version_2"}
                old["items"] = old.pop("items")
                dump(f, old,
                     "# Superseded by version 2 (spec 0007). Kept so runs that used this\n"
                     "# version still resolve; never edit a published version in place.\n")
                dump(CONSTRUCTS / f"{cid}_v2.yaml", live,
                     f"# Construct: versioned, append-only. Version 2 applies the "
                     f"{REVIEW_DATE} library review (spec 0007).\n")
        else:
            same = all(live.get(k) == c.get(k) for k in
                       ("verification_status", "source_url", "citation", "questionnaire"))
            (plan["unchanged"] if same and "review" in c else plan["in_place"]).append(cid)
            if write:
                dump(f, live,
                     "# Construct: versioned, append-only. Edits create a NEW version "
                     "(see registries rule).\n")

    # A construct that no longer needs a new version must not leave one behind:
    # a stale v2 would keep winning in the picker (newest version per slug).
    stale = [
        p for p in CONSTRUCTS.glob("*_v2.yaml")
        if yaml.safe_load(p.read_text())["construct_id"] not in plan["new_version"]
    ]
    for p in stale:
        print(f"stale version 2 file (no longer needed): {p.name}")
        if write:
            p.unlink()

    print("\n".join(plan["detail"]) if plan["detail"] else "  (no item-level changes)")
    live_total = len(plan["new_version"]) + len(plan["in_place"]) + len(plan["unchanged"])
    print(f"\nitem-level changes  : {len(plan['detail'])}")
    print(f"new version 2 files : {len(plan['new_version'])}")
    print(f"in-place metadata   : {len(plan['in_place'])}")
    print(f"already current     : {len(plan['unchanged'])}")
    print(f"\nlive constructs     : {live_total}")
    print(f"  verified          : {live_total - len(PENDING)}")
    print(f"  needs_verification: {len(PENDING)}")
    for cid in sorted(PENDING):
        print(f"      {cid}")
    if not write:
        print("\nDry run. Re-run with --write to apply.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
